import json
import logging
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

import src.agents.report_agent as report_agent_module
from src.agents.report_agent import (
    ReportAgent,
    ReportAnswer,
    ReportPayload,
    ReportSection,
    ReportTrust,
)
from src.tools.base import ToolContext
from src.tools.document_tools import CreateWorkbookTool
from src.workflows.workbook_models import (
    WorkbookChartSpec,
    WorkbookFinancialRow,
    WorkbookSheetSpec,
    WorkbookSpec,
    WorkbookTable,
)


def _make_agent() -> ReportAgent:
    return ReportAgent(tools=None)  # type: ignore[arg-type]


def _make_payload() -> ReportPayload:
    return ReportPayload(
        answer=ReportAnswer(
            title="Weekly Financial Check-In",
            summary="This week financial view.",
            sections=[
                ReportSection(label="Key Finding", items=["North America revenue reached $35.0M this week."]),
                ReportSection(label="Business Implications", items=["AWS cost increased to $4.5M."]),
                ReportSection(label="Recommended Actions", items=["Protect runway and review spend."]),
            ],
        ),
        trust=ReportTrust(
            confidence="medium",
            confidence_score=0.72,
            assumptions=[],
            open_questions=[],
            data_quality="high",
        ),
        sources=[],
    )


def test_financial_rows_use_normalized_periods_taxonomy_and_provenance() -> None:
    agent = _make_agent()
    rows = agent._build_financial_rows(
        task_input="Generate this week's financial workbook based off prior week performance",
        company_state={
            "revenue_segmentation": {"north america": 35_000_000},
            "cost_structure": {"aws": 4_500_000},
            "capital_position": {"cash at bank": 15_200_000},
        },
        metrics=[],
        ceo_id=None,
        current_interaction_id=None,
        session_history=[],
        retrieval=[
            {
                "title": "Weekly Finance Notes",
                "content": (
                    "Prior week AWS cost budget $4.1M actual $4.5M forecast $4.7M variance $0.4M.\n"
                    "Q1 2026 North America revenue budget $34.0M actual $35.0M forecast $36.0M variance $1.0M."
                ),
            }
        ],
    )

    metrics = {row.metric: row for row in rows}

    assert "AWS cost" in metrics
    assert metrics["AWS cost"].period == "Prior Week"
    assert metrics["AWS cost"].source_type == "retrieved_document"
    assert metrics["AWS cost"].source_ref == "Weekly Finance Notes"

    assert "North America revenue" in metrics
    assert metrics["North America revenue"].period == "Q1 2026"
    assert metrics["North America revenue"].actual == 35_000_000

    assert "Cash at bank" in metrics
    assert metrics["Cash at bank"].period == "Current Week"
    assert metrics["Cash at bank"].source_ref == "CompanyState.capital_position.cash at bank"


def test_finance_workbook_spec_exposes_tabs_and_row_provenance() -> None:
    agent = _make_agent()
    spec = agent._to_workbook_spec(
        task_input="Generate a weekly financial analysis workbook with charts and variance",
        payload=_make_payload(),
        company_state={
            "revenue_segmentation": {"north america": 35_000_000},
            "cost_structure": {"aws": 4_500_000},
            "capital_position": {"cash at bank": 15_200_000},
        },
        retrieval=[],
    )

    sheet_names = [sheet.name for sheet in spec.sheets]
    assert sheet_names == ["Summary", "Model", "Variance", "Forecast", "Charts"]

    model_sheet = next(sheet for sheet in spec.sheets if sheet.name == "Model")
    model_table = model_sheet.tables[0]
    assert model_table.columns[-1] == "Source"
    assert len(model_table.row_provenance) == len(model_table.rows)
    assert model_table.row_provenance[0]["source_type"] in {
        "company_state",
        "retrieved_document",
        "derived_metric",
        "fallback",
    }


def test_finance_workbook_builds_week_over_week_comparison_table() -> None:
    agent = _make_agent()
    spec = agent._to_workbook_spec(
        task_input="Generate this week's financial analysis workbook based off the previous week",
        payload=_make_payload(),
        company_state={
            "revenue_segmentation": {"north america": 35_000_000},
            "cost_structure": {"aws": 4_500_000},
            "capital_position": {"cash at bank": 15_200_000},
        },
        retrieval=[
            {
                "title": "Prior Week Finance Notes",
                "content": (
                    "Prior week North America revenue budget $33.0M actual $34.0M forecast $35.0M variance $1.0M.\n"
                    "Prior week AWS cost budget $4.0M actual $4.2M forecast $4.4M variance $0.2M."
                ),
            }
        ],
    )

    summary_sheet = next(sheet for sheet in spec.sheets if sheet.name == "Summary")
    assert any(metric.label == "Period Delta" for metric in summary_sheet.metrics)

    variance_sheet = next(sheet for sheet in spec.sheets if sheet.name == "Variance")
    comparison_table = next(table for table in variance_sheet.tables if table.title == "Period Comparison")
    assert comparison_table.columns == [
        "Metric",
        "Prior Period",
        "Current Period",
        "Prior Actual",
        "Current Actual",
        "Delta",
        "Delta %",
    ]
    revenue_row = next(row for row in comparison_table.rows if row[0] == "North America revenue")
    assert revenue_row[1] == "Prior Week"
    assert revenue_row[2] == "Current Week"
    assert revenue_row[5] == "$1.0M"
    assert revenue_row[6].endswith("%")

    charts_sheet = next(sheet for sheet in spec.sheets if sheet.name == "Charts")
    assert any(table.title == "Period Comparison Data" for table in charts_sheet.tables)
    assert any(chart.title == "Period Delta by Metric" for chart in charts_sheet.chart_specs)


def test_finance_workbook_uses_dated_historical_artifact_for_prior_month(monkeypatch) -> None:
    agent = _make_agent()
    now = datetime.now()
    prior_month = now.month - 1 or 12
    prior_month_year = now.year if now.month > 1 else now.year - 1

    historical_spec = WorkbookSpec(
        workbook_title="Monthly Finance Review",
        sheets=[
            WorkbookSheetSpec(
                name="Model",
                kind="model",
                financial_rows=[
                    WorkbookFinancialRow(
                        period="Current Month",
                        metric="North America revenue",
                        budget=31_000_000,
                        actual=32_000_000,
                        variance=1_000_000,
                        forecast=33_000_000,
                    )
                ],
            )
        ],
    )

    monkeypatch.setattr(
        report_agent_module,
        "read_stage_artifact",
        lambda interaction_id, ceo_id, stage: json.dumps(historical_spec.model_dump()) if interaction_id == 77 else "",
    )

    spec = agent._to_workbook_spec(
        task_input="Generate this month's financial analysis workbook based off last month",
        payload=_make_payload(),
        company_state={
            "revenue_segmentation": {"north america": 35_000_000},
            "cost_structure": {"aws": 4_500_000},
            "capital_position": {"cash at bank": 15_200_000},
        },
        ceo_id="ceo_test",
        current_interaction_id=90,
        session_history=[
            {"id": 77, "timestamp": f"{prior_month_year:04d}-{prior_month:02d}-10T09:00:00", "query": "Last month check-in"},
            {"id": 66, "timestamp": f"{now.year:04d}-{now.month:02d}-02T09:00:00", "query": "Current month check-in"},
        ],
        retrieval=[],
    )

    variance_sheet = next(sheet for sheet in spec.sheets if sheet.name == "Variance")
    comparison_table = next(table for table in variance_sheet.tables if table.title == "Period Comparison")
    revenue_row = next(row for row in comparison_table.rows if row[0] == "North America revenue")
    assert revenue_row[1] == "Prior Month"
    assert revenue_row[2] == "Current Month"
    assert revenue_row[5] == "$3.0M"

    model_sheet = next(sheet for sheet in spec.sheets if sheet.name == "Model")
    assert any(item["source_type"] == "historical_artifact" for item in model_sheet.tables[0].row_provenance)


def test_historical_selection_prefers_explicit_period_metadata(monkeypatch) -> None:
    agent = _make_agent()
    now = datetime.now()

    monkeypatch.setattr(
        report_agent_module,
        "read_stage_artifact_metadata",
        lambda interaction_id, ceo_id, stage: (
            {"period_coverage": {"periods": ["Prior Quarter"], "comparison_pairs": [{"prior": "Prior Quarter", "current": "Current Quarter"}]}}
            if interaction_id == 88
            else {}
        ),
    )

    selected = agent._select_historical_interaction_for_period(
        session_history=[
            {"id": 88, "timestamp": f"{now.year:04d}-{now.month:02d}-01T09:00:00", "query": "Labeled prior quarter workbook"},
            {"id": 77, "timestamp": f"{now.year:04d}-01-10T09:00:00", "query": "Unlabeled older workbook"},
        ],
        target_period="Prior Quarter",
        current_interaction_id=99,
        current_timestamp=now,
        ceo_id="ceo_test",
    )

    assert selected is not None
    assert selected["id"] == 88


def test_create_workbook_only_adds_charts_for_numeric_tables(tmp_path: Path) -> None:
    tool = CreateWorkbookTool()
    context = ToolContext(interaction_id=99, ceo_id="ceo_test")
    spec = WorkbookSpec(
        workbook_title="Finance Workbook",
        sheets=[
            WorkbookSheetSpec(
                name="Model",
                kind="model",
                tables=[
                    WorkbookTable(
                        title="Numeric Table",
                        columns=["Metric", "Actual"],
                        rows=[["Revenue", "35.0M"], ["AWS cost", "4.5M"]],
                    ),
                    WorkbookTable(
                        title="Text Table",
                        columns=["Metric", "Actual"],
                        rows=[["Revenue", "strong"], ["AWS cost", "elevated"]],
                    ),
                ],
                chart_specs=[
                    WorkbookChartSpec(
                        title="Numeric Chart",
                        chart_type="bar",
                        x_axis="Metric",
                        y_axis="Actual",
                        series_label="Actual",
                        source_sheet="Model",
                        source_table="Numeric Table",
                    ),
                    WorkbookChartSpec(
                        title="Text Chart",
                        chart_type="bar",
                        x_axis="Metric",
                        y_axis="Actual",
                        series_label="Actual",
                        source_sheet="Model",
                        source_table="Text Table",
                    ),
                ],
            )
        ],
    )

    result = tool.invoke(
        context,
        output_path=str(tmp_path / "finance.xlsx"),
        workbook_spec=spec.model_dump(),
    )

    assert result.success is True
    assert result.metadata["preview_metadata"]["period_coverage"]["comparison_pairs"] == []
    workbook = load_workbook(tmp_path / "finance.xlsx")
    model_sheet = workbook["Model"]
    assert len(model_sheet._charts) == 1
    assert model_sheet._charts[0].title.tx.rich.p[0].r[0].t == "Numeric Chart"


def test_create_workbook_accepts_top_level_fields(tmp_path: Path) -> None:
    tool = CreateWorkbookTool()
    context = ToolContext(interaction_id=100, ceo_id="ceo_test")

    result = tool.invoke(
        context,
        output_path=str(tmp_path / "finance-top-level.xlsx"),
        workbook_title="Finance Workbook",
        sheets=[
            {
                "name": "Model",
                "kind": "model",
                "tables": [
                    {
                        "title": "Numeric Table",
                        "columns": ["Metric", "Actual"],
                        "rows": [["Revenue", "35.0M"]],
                    }
                ],
            }
        ],
        template_id="finance_workbook_v1",
        theme_id="board_formal",
    )

    assert result.success is True
    assert result.metadata["view_model"]["title"] == "Finance Workbook"
    workbook = load_workbook(tmp_path / "finance-top-level.xlsx")
    assert workbook.sheetnames == ["Model"]


def test_create_workbook_logs_invalid_payload_keys(caplog, tmp_path: Path) -> None:
    tool = CreateWorkbookTool()
    context = ToolContext(interaction_id=101, ceo_id="ceo_test", stage="synthesizer")

    with caplog.at_level(logging.WARNING):
        result = tool.invoke(
            context,
            output_path=str(tmp_path / "invalid.xlsx"),
            workbook_spec={"title": "Wrong Key", "sheets": []},
        )

    assert result.success is False
    assert "Invalid workbook payload" in (result.error or "")
    assert "Invalid workbook payload for create_workbook" in caplog.text
    assert "payload_keys=['sheets', 'title']" in caplog.text
    assert "kwargs_keys=['output_path', 'workbook_spec']" in caplog.text
