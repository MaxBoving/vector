from pathlib import Path

from openpyxl import load_workbook

from src.presentation.render_xlsx import build_workbook_period_metadata, render_xlsx_workbook
from src.workflows.workbook_models import WorkbookChartSpec, WorkbookSheetSpec, WorkbookSpec, WorkbookTable


def test_render_xlsx_workbook_writes_file_and_view_metadata(tmp_path: Path) -> None:
    path = tmp_path / "analysis.xlsx"
    spec = WorkbookSpec(
        workbook_title="Finance Workbook",
        metadata={"template_id": "finance_workbook_v1", "theme_id": "board_formal"},
        sheets=[
            WorkbookSheetSpec(
                name="Model",
                kind="model",
                tables=[
                    WorkbookTable(
                        title="Numeric Table",
                        columns=["Metric", "Actual"],
                        rows=[["Revenue", "35.0M"], ["AWS cost", "4.5M"]],
                    )
                ],
                chart_specs=[
                    WorkbookChartSpec(
                        title="Numeric Chart",
                        chart_type="bar",
                        x_axis="Metric",
                        y_axis="Actual",
                        series_label="Actual",
                        source_sheet="Model",
                        source_table="Numeric Table",
                    )
                ],
            )
        ],
    )

    result = render_xlsx_workbook(path=path, spec=spec, artifact_id="interaction:1:analysis_xlsx")

    assert path.exists()
    assert result["preview_format"] == "json"
    assert result["preview_metadata"]["template_id"] == "finance_workbook_v1"
    assert result["preview_metadata"]["theme_id"] == "board_formal"
    assert result["view_model"]["metadata"]["template_id"] == "finance_workbook_v1"
    workbook = load_workbook(path)
    worksheet = workbook["Model"]
    assert "Model" in workbook.sheetnames
    assert len(worksheet._charts) == 1
    assert worksheet["A1"].font.color.rgb in {"001E3A5F", "1E3A5F"}
    assert worksheet["A1"].fill.fgColor.rgb in {"00F3E8D8", "F3E8D8"}
    assert worksheet["A4"].fill.fgColor.rgb in {"001E3A5F", "1E3A5F"}


def test_build_workbook_period_metadata_extracts_comparison_pairs() -> None:
    spec = WorkbookSpec(
        workbook_title="Finance Workbook",
        sheets=[
            WorkbookSheetSpec(
                name="Variance",
                kind="variance",
                tables=[
                    WorkbookTable(
                        title="Period Comparison",
                        columns=["Metric", "Prior Period", "Current Period"],
                        rows=[["Revenue", "Prior Quarter", "Current Quarter"]],
                    )
                ],
            )
        ],
    )

    metadata = build_workbook_period_metadata(spec)
    assert metadata["period_coverage"]["comparison_pairs"] == [{"prior": "Prior Quarter", "current": "Current Quarter"}]


def test_render_xlsx_workbook_uses_template_tab_order(tmp_path: Path) -> None:
    path = tmp_path / "template-ordered-workbook.xlsx"
    spec = WorkbookSpec(
        workbook_title="Finance Workbook",
        metadata={"template_id": "finance_workbook_v1", "theme_id": "board_formal"},
        sheets=[
            WorkbookSheetSpec(name="Charts", kind="charts"),
            WorkbookSheetSpec(name="Summary", kind="summary"),
            WorkbookSheetSpec(name="Model", kind="model"),
        ],
    )

    render_xlsx_workbook(path=path, spec=spec, artifact_id="interaction:1:analysis_xlsx")

    workbook = load_workbook(path)
    assert workbook.sheetnames[:3] == ["Summary", "Model", "Charts"]


def test_render_xlsx_workbook_uses_template_tab_content_policy(tmp_path: Path) -> None:
    path = tmp_path / "template-content-policy.xlsx"
    spec = WorkbookSpec(
        workbook_title="Finance Workbook",
        metadata={"template_id": "finance_workbook_v1", "theme_id": "board_formal"},
        sheets=[
            WorkbookSheetSpec(
                name="Summary",
                kind="summary",
                tables=[WorkbookTable(title="Executive Summary", columns=["Section", "Detail"], rows=[["A", "B"]])],
                chart_specs=[
                    WorkbookChartSpec(
                        title="Should Not Render",
                        chart_type="bar",
                        x_axis="Section",
                        y_axis="Detail",
                        series_label="Detail",
                        source_sheet="Summary",
                        source_table="Executive Summary",
                    )
                ],
            ),
            WorkbookSheetSpec(
                name="Charts",
                kind="charts",
                tables=[WorkbookTable(title="Chart Table", columns=["Metric", "Actual"], rows=[["Revenue", "35.0M"]])],
                chart_specs=[
                    WorkbookChartSpec(
                        title="Should Render",
                        chart_type="bar",
                        x_axis="Metric",
                        y_axis="Actual",
                        series_label="Actual",
                        source_sheet="Charts",
                        source_table="Chart Table",
                    )
                ],
            ),
        ],
    )

    render_xlsx_workbook(path=path, spec=spec, artifact_id="interaction:1:analysis_xlsx")

    workbook = load_workbook(path)
    assert len(workbook["Summary"]._charts) == 0
    assert len(workbook["Charts"]._charts) == 1
