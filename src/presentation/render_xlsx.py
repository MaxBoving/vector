from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill

from src.presentation import get_workbook_template, resolve_brand_theme
from src.presentation.artifact_contracts import DEFAULT_WORKBOOK_TEMPLATE_ID, normalize_workbook_spec
from src.presentation.render_qa import qa_check_xlsx
from src.workflows.workbook_models import WorkbookSpec, workbook_spec_to_view_model


def _sanitize_sheet_name(name: str) -> str:
    cleaned = "".join(character for character in name if character not in "\\/*?:[]")
    return (cleaned or "Sheet")[:31]


def _parse_numeric_cell(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    if not isinstance(value, str):
        return None
    normalized = value.replace("$", "").replace(",", "").strip()
    multiplier = 1.0
    if normalized.endswith("%"):
        normalized = normalized[:-1]
    if normalized.endswith("M"):
        multiplier = 1_000_000.0
        normalized = normalized[:-1]
    elif normalized.endswith("K"):
        multiplier = 1_000.0
        normalized = normalized[:-1]
    try:
        return float(normalized) * multiplier
    except ValueError:
        return None


def _find_chart_columns(columns: list[str], y_axis: str) -> tuple[int, int] | None:
    lowered_columns = [str(column).strip().lower() for column in columns]
    category_candidates = ["metric", "period", "section", "label"]
    category_index = next((lowered_columns.index(name) + 1 for name in category_candidates if name in lowered_columns), 1)
    value_key = str(y_axis or "actual").strip().lower()
    if value_key in lowered_columns:
        return category_index, lowered_columns.index(value_key) + 1
    numeric_candidates = ["actual", "forecast", "budget", "variance", "value"]
    matched = next((name for name in numeric_candidates if name in lowered_columns), None)
    if not matched:
        return None
    return category_index, lowered_columns.index(matched) + 1


def _table_has_numeric_series(worksheet: Any, *, data_start_row: int, data_end_row: int, value_index: int) -> bool:
    if data_end_row < data_start_row:
        return False
    for row_index in range(data_start_row, data_end_row + 1):
        value = worksheet.cell(row=row_index, column=value_index).value
        if isinstance(value, (int, float)):
            return True
    return False


def build_workbook_period_metadata(spec: WorkbookSpec) -> dict[str, Any]:
    periods: list[str] = []
    comparison_pairs: list[dict[str, str]] = []
    for sheet in spec.sheets or []:
        for row in sheet.financial_rows or []:
            if row.period and row.period not in periods:
                periods.append(row.period)
        for table in sheet.tables or []:
            if table.title != "Period Comparison":
                continue
            for row in table.rows or []:
                if len(row) < 3:
                    continue
                prior_period = str(row[1])
                current_period = str(row[2])
                pair = {"prior": prior_period, "current": current_period}
                if pair not in comparison_pairs:
                    comparison_pairs.append(pair)
                for period in (prior_period, current_period):
                    if period and period not in periods:
                        periods.append(period)
    return {
        "artifact_role": "financial_analysis_workbook",
        "period_coverage": {
            "periods": periods,
            "comparison_pairs": comparison_pairs,
            "has_comparison": bool(comparison_pairs),
        },
    }


def _template_tab_rule(workbook_template: Any, sheet_name: str, sheet_kind: str) -> Any | None:
    for tab_spec in workbook_template.tab_specs:
        if tab_spec.name == sheet_name or tab_spec.kind == sheet_kind:
            return tab_spec
    return None


def render_xlsx_workbook(*, path: Path, spec: WorkbookSpec, artifact_id: str) -> dict[str, object]:
    spec = normalize_workbook_spec(spec)
    theme = resolve_brand_theme(spec.metadata.get("theme_id"))
    workbook_template = get_workbook_template(str(spec.metadata.get("template_id") or DEFAULT_WORKBOOK_TEMPLATE_ID))
    tab_rank = {tab_name: index for index, tab_name in enumerate(workbook_template.tab_order)}
    ordered_sheets = sorted(
        spec.sheets or [],
        key=lambda sheet: (tab_rank.get(str(sheet.name), len(tab_rank) + 1), str(sheet.name)),
    )
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    table_locations: dict[tuple[str, str], dict[str, Any]] = {}

    for sheet_index, sheet_data in enumerate(ordered_sheets):
        tab_rule = _template_tab_rule(workbook_template, str(sheet_data.name), str(sheet_data.kind))
        worksheet = workbook.create_sheet(_sanitize_sheet_name(str(sheet_data.name or f"Sheet {sheet_index + 1}")))
        row_pointer = 1
        worksheet["A1"] = spec.workbook_title
        worksheet["A1"].font = Font(
            size=14,
            bold=True,
            color=theme.colors.primary.replace("#", ""),
            name=theme.typography.heading_family,
        )
        worksheet["A1"].fill = PatternFill("solid", fgColor=theme.tables.emphasis_background.replace("#", ""))
        row_pointer += 2

        metrics = list(sheet_data.metrics or []) if (tab_rule.allow_metrics if tab_rule else True) else []
        if metrics:
            worksheet.cell(row=row_pointer, column=1, value="Metric").font = Font(
                bold=True,
                color=theme.tables.header_text.replace("#", ""),
                name=theme.typography.body_family,
            )
            worksheet.cell(row=row_pointer, column=1).fill = PatternFill(
                "solid", fgColor=theme.tables.header_background.replace("#", "")
            )
            worksheet.cell(row=row_pointer, column=2, value="Value").font = Font(
                bold=True,
                color=theme.tables.header_text.replace("#", ""),
                name=theme.typography.body_family,
            )
            worksheet.cell(row=row_pointer, column=2).fill = PatternFill(
                "solid", fgColor=theme.tables.header_background.replace("#", "")
            )
            row_pointer += 1
            for metric in metrics:
                worksheet.cell(row=row_pointer, column=1, value=str(metric.label))
                worksheet.cell(row=row_pointer, column=2, value=str(metric.value))
                fill_color = (
                    theme.tables.row_even_background if row_pointer % 2 == 0 else theme.tables.row_odd_background
                ).replace("#", "")
                worksheet.cell(row=row_pointer, column=1).fill = PatternFill("solid", fgColor=fill_color)
                worksheet.cell(row=row_pointer, column=2).fill = PatternFill("solid", fgColor=fill_color)
                row_pointer += 1
            row_pointer += 1

        tables = list(sheet_data.tables or []) if (tab_rule.allow_tables if tab_rule else True) else []
        for table in tables:
            title = str(table.title or "Table")
            columns = [str(column) for column in (table.columns or [])]
            rows = [[str(cell) for cell in row] for row in (table.rows or [])]
            worksheet.cell(row=row_pointer, column=1, value=title).font = Font(
                bold=True,
                color=theme.colors.secondary.replace("#", ""),
                name=theme.typography.heading_family,
            )
            row_pointer += 1
            for column_index, column_name in enumerate(columns, start=1):
                worksheet.cell(row=row_pointer, column=column_index, value=column_name).font = Font(
                    bold=True,
                    color=theme.tables.header_text.replace("#", ""),
                    name=theme.typography.body_family,
                )
                worksheet.cell(row=row_pointer, column=column_index).fill = PatternFill(
                    "solid", fgColor=theme.tables.header_background.replace("#", "")
                )
            row_pointer += 1
            data_start_row = row_pointer
            for row in rows:
                fill_color = (
                    theme.tables.row_even_background if row_pointer % 2 == 0 else theme.tables.row_odd_background
                ).replace("#", "")
                for column_index, cell in enumerate(row, start=1):
                    numeric_value = _parse_numeric_cell(cell)
                    worksheet.cell(row=row_pointer, column=column_index, value=numeric_value if numeric_value is not None else cell)
                    worksheet.cell(row=row_pointer, column=column_index).fill = PatternFill("solid", fgColor=fill_color)
                row_pointer += 1
            table_locations[(sheet_data.name, title)] = {
                "data_start_row": data_start_row,
                "data_end_row": max(data_start_row, row_pointer - 1),
                "columns": columns,
            }
            row_pointer += 2

        chart_specs = list(sheet_data.chart_specs or []) if (tab_rule.allow_charts if tab_rule else True) else []
        for chart_spec in chart_specs:
            source_sheet = chart_spec.source_sheet or sheet_data.name
            source_table = chart_spec.source_table
            if not source_table:
                continue
            table_location = table_locations.get((source_sheet, source_table))
            if not table_location:
                continue
            data_start_row = table_location["data_start_row"]
            data_end_row = table_location["data_end_row"]
            if data_end_row < data_start_row:
                continue
            column_indices = _find_chart_columns(table_location["columns"], str(chart_spec.y_axis or "actual"))
            if not column_indices:
                continue
            category_index, value_index = column_indices
            source_worksheet = workbook[_sanitize_sheet_name(source_sheet)]
            if not _table_has_numeric_series(
                source_worksheet,
                data_start_row=data_start_row,
                data_end_row=data_end_row,
                value_index=value_index,
            ):
                continue
            chart = BarChart()
            chart.title = str(chart_spec.title or "Chart")
            chart.style = 10
            chart.y_axis.title = str(chart_spec.y_axis or "Value")
            chart.x_axis.title = str(chart_spec.x_axis or "Category")
            values = Reference(source_worksheet, min_col=value_index, min_row=data_start_row, max_row=data_end_row)
            categories = Reference(source_worksheet, min_col=category_index, min_row=data_start_row, max_row=data_end_row)
            chart.add_data(values, titles_from_data=False)
            chart.set_categories(categories)
            worksheet.add_chart(chart, f"F{max(4, row_pointer)}")
            row_pointer += 14

        if tab_rule and not tab_rule.allow_pivots:
            pass

        worksheet.freeze_panes = "A3"
        worksheet.column_dimensions["A"].width = 26
        worksheet.column_dimensions["B"].width = 18
        worksheet.column_dimensions["C"].width = 22
        worksheet.column_dimensions["D"].width = 48

    workbook.save(path)
    qa_report = qa_check_xlsx(path, spec)
    return {
        "preview_content": json.dumps(spec.model_dump(), indent=2),
        "preview_format": "json",
        "preview_metadata": {
            **spec.metadata,
            **build_workbook_period_metadata(spec),
        },
        "view_model": workbook_spec_to_view_model(spec, artifact_id=artifact_id),
        "qa_report": qa_report.model_dump(),
    }
